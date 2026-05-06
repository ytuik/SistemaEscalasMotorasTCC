"""
Planilha_Service - Importação de avaliações a partir de modelo criado no Template_Service

Estrutura esperada do arquivo
-----------------------
Uma linha por item respondido. Avaliações do mesmo paciente na mesma
data são agrupadas automaticamente.
Nas paginas especificas de cada escala, os itens devem ser listados na ordem numérica, e com o resultado preenchido. 

Aba - Informações da Avaliação:
    C4  — paciente_nome
    C5  — data_nascimento    (DD-MM-AAAA)
    C6  — fisioterapeuta_crefito
    C7  — data_avaliacao     (DD-MM-AAAA)
    C8  — observacoes        (opcional)
                          
Abas específicas de cada escala: 
    Coluna A — numero_item   (preenchido pelo template)
    Coluna B — descricao     (preenchido pelo template)
    Coluna C — pontuacao     (preenchido pelo fisioterapeuta)
    Coluna D — pontuacao_max (preenchido pelo template)
    Linha de TOTAL é ignorada automaticamente.
    
    Abas sem nenhuma pontuação preenchida são ignoradas silenciosamente
    — o fisioterapeuta simplesmente deixa em branco as escalas não aplicadas.
"""

import logging

from openpyxl import load_workbook
from sqlalchemy.orm import Session
from app.exceptions import PlanilhaInvalidaError, PlanilhaNaoEncontradaError
from app.models import Escala
from app.models.dto.aplicacao_input_dto import AplicacaoInputDTO
from app.models.dto.nova_avaliacao_dto import NovaAvaliacaoDTO
from app.models.dto.resposta_input_dto import RespostaInputDTO
from app.services import obter_fisioterapeuta_por_crefito, registrar_avaliacao, obter_escalas
from app.utils.string_parser import parse_data

ABA_INFO = "Informações da Avaliação"
LINHA_PACIENTE_NOME = 4
LINHA_DATA_NASCIMENTO = 5
LINHA_FISIOTERAPEUTA_CREFITO = 6
LINHA_DATA_AVALIACAO = 7
LINHA_OBSERVACOES = 8
COLUNA_VALOR = 3

logger = logging.getLogger(__name__)

def importar_planilha_avaliacao(session: Session, caminho_arquivo: str) -> dict:
    """Importa avaliações a partir de um arquivo XLSX e as registra no banco de dados.

    O arquivo deve seguir o formato especificado na documentação da função.

    Parâmetros
    ----------
    session : Session
        Sessão SQLAlchemy ativa.
    caminho_arquivo : str
        Caminho para o arquivo XLSX a ser importado.

    Retorna
    -------
    dict
        Um dicionário contendo o resultado da importação, incluindo o número de avaliações importadas e detalhes de cada avaliação.

    """
    workbook = None
    # Carregar o arquivo XLSX usando OpenPyXL
    try:
        workbook = load_workbook(filename=caminho_arquivo, data_only=True)
    
        if ABA_INFO not in workbook.sheetnames:
            raise PlanilhaInvalidaError(f"Aba '{ABA_INFO}' não encontrada na planilha.")
        
        info = _ler_aba_informacoes(workbook[ABA_INFO], caminho_arquivo)
        
        fisioterapeuta = obter_fisioterapeuta_por_crefito(session, info["fisioterapeuta_crefito"])
        if not fisioterapeuta:
            raise PlanilhaInvalidaError(f"Fisioterapeuta com CREFITO '{info['fisioterapeuta_crefito']}' não encontrado no banco de dados.")
        
        escalas_respostas = _ler_abas_escalas(workbook, session)
        
        if not escalas_respostas:
            raise PlanilhaInvalidaError(f"Nenhuma escala com respostas preenchidas encontrada na planilha '{caminho_arquivo}'.")
        
        payload = NovaAvaliacaoDTO(
            nome_paciente=info["paciente_nome"],
            data_nascimento=info["data_nascimento"],
            fisioterapeuta_id=fisioterapeuta.id,
            data_avaliacao=info["data_avaliacao"],
            observacoes=info["observacoes"],
            aplicacoes=escalas_respostas
        )
        avaliacao = registrar_avaliacao(session=session, payload=payload)
            
        return {
            "status": "ok",
            "paciente": info["paciente_nome"],
            "data_avaliacao": info["data_avaliacao"].isoformat(),
            "avaliacao_id": avaliacao.id,
            "escalas_importadas": [escala.id_escala for escala in escalas_respostas],
            "mensagem": None
        }
            
    except FileNotFoundError:
        raise PlanilhaNaoEncontradaError(f"Arquivo não encontrado: {caminho_arquivo}")
    except (PlanilhaInvalidaError, PlanilhaNaoEncontradaError, Exception) as e:
        return {
            "status": "erro",
            "paciente": info["paciente_nome"],
            "data_avaliacao": info["data_avaliacao"].isoformat(),
            "mensagem": str(e)
        }
    except Exception as e:
        logger.error(f"Erro inesperado ao importar planilha {caminho_arquivo}: {e}", exc_info=True)
        return {
            "status": "erro",
            "mensagem": f"Erro interno ao processar a planilha: {str(e)}"
        }
    finally:
        if workbook:
            workbook.close()
        
def _ler_aba_informacoes(workbook, filepath: str) -> dict:
    """Lê a aba 'Informações da Avaliação' e extrai os dados necessários."""
    
    def get_valor(linha):
        valor = workbook.cell(row=linha, column=COLUNA_VALOR).value
        return str(valor).strip() if valor is not None else None
    
    paciente_nome = get_valor(LINHA_PACIENTE_NOME)
    data_nascimento_str = get_valor(LINHA_DATA_NASCIMENTO)
    fisioterapeuta_crefito = get_valor(LINHA_FISIOTERAPEUTA_CREFITO)
    data_avaliacao_str = get_valor(LINHA_DATA_AVALIACAO)
    observacoes = get_valor(LINHA_OBSERVACOES) or ""
    
    erros = []
    if not paciente_nome:
        erros.append("Campo 'paciente_nome' está vazio.")
    if not data_nascimento_str:
        erros.append("Campo 'data_nascimento' está vazio.")
    if not fisioterapeuta_crefito:
        erros.append("Campo 'fisioterapeuta_crefito' está vazio.")
    if not data_avaliacao_str:
        erros.append("Campo 'data_avaliacao' está vazio.")
    if erros:
        raise PlanilhaInvalidaError(f"Erros encontrados na aba '{ABA_INFO}' do arquivo '{filepath}': " + "; ".join(erros))
    
    try:
        data_nascimento = parse_data(data_nascimento_str)
    except ValueError:
        raise PlanilhaInvalidaError(f"Formato inválido para 'data_nascimento' na aba '{ABA_INFO}' do arquivo '{filepath}'.")
    
    try:
        data_avaliacao = parse_data(data_avaliacao_str)
    except ValueError:
        raise PlanilhaInvalidaError(f"Formato inválido para 'data_avaliacao' na aba '{ABA_INFO}' do arquivo '{filepath}'.")
    
    return {
        "paciente_nome": paciente_nome,
        "data_nascimento": data_nascimento,
        "fisioterapeuta_crefito": fisioterapeuta_crefito,
        "data_avaliacao": data_avaliacao,
        "observacoes": observacoes
    }
    
def _ler_abas_escalas(workbook, session: Session) -> list[AplicacaoInputDTO]:
    """
    Itera pelas abas da planilha para identificar as escalas aplicadas e extrair as respostas dos itens.
    
    Abas sem nenhuma pontuação preenchida são ignoradas silenciosamente.
    """
    
    escalas_respostas = []
    todas_escalas = obter_escalas(session)

    for nome_aba in workbook.sheetnames:
        if nome_aba == ABA_INFO:
            continue

        aba = workbook[nome_aba]
        escala = _match_escala(nome_aba, todas_escalas)
        if not escala:
            print(f"Aba '{nome_aba}' não corresponde a nenhuma escala conhecida. Ignorando esta aba.")
            continue
        
        respostas = _ler_respostas_escalas(aba)
        if not respostas:
            print(f"Aba '{nome_aba}' não tem respostas preenchidas. Ignorando esta aba.")
            continue
        
        escalas_respostas.append(AplicacaoInputDTO(
            id_escala=escala.id,
            respostas=respostas
        ))

    return escalas_respostas

def _match_escala(nome_aba: str, escalas:list) -> Escala | None:
    """
    Tenta encontrar uma escala no banco de dados que corresponda ao nome da aba da planilha.
    
    Estratégia: verifica se o nome da aba aparece dentro do nome
    completo da escala no banco (case insensitive). Funciona porque
    todas as abreviações usadas no template (Berg, TUG, DGI, DASH...)
    estão contidas nos nomes completos cadastrados.
    """
    nome_aba_normalizado = nome_aba.strip().lower()
    for escala in escalas:
        if nome_aba_normalizado in escala.nome.strip().lower():
            return escala
    return None

def _ler_respostas_escalas(aba) -> list[RespostaInputDTO]:
    """    
    Lê as respostas de uma aba de escala.

    Começa na linha 4 (primeira linha de item) e para quando encontra
    a linha de TOTAL ou uma linha sem número de item na coluna A.
    Ignora linhas onde a pontuação (coluna C) não foi preenchida.

    Retorna uma lista de RespostaInputDTO apenas com itens preenchidos.
    
    """
    
    respostas = []
    for row in aba.iter_rows(min_row=4, values_only=True):
        numero_item = row[0] # Coluna A
        pontuacao = row[2]   # Coluna C
        
        if numero_item is None or str(numero_item).strip().lower() == "total":
            break
        if not str(numero_item).strip().isdigit():
            continue
        if pontuacao is None or str(pontuacao).strip() == "":
            continue
        
        try:
            respostas.append(RespostaInputDTO(
                numero_item=int(numero_item),
                pontuacao=int(float(pontuacao))
            ))
        except (ValueError, TypeError):
            print(f"Valor de pontuação inválido para item {numero_item} na aba '{aba.title}'. Ignorando este item.")
            continue
    
    return respostas
        