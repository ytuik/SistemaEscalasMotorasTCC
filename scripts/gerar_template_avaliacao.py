"""
Gera o template Excel padrão de avaliação.
O arquivo é salvo em data/template_avaliacao.xlsx e serve como
base para todas as avaliações — nunca deve ser editado diretamente.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)

from scripts.utils.escalas import get_escalas

AZUL_HEADER = "1F4E79"
AZUL_CLARO  = "BDD7EE"
CINZA_LINHA = "F2F2F2"
VERDE_INPUT = "E2EFDA"
AMARELO_OBS = "FFF2CC"
BRANCO      = "FFFFFF"
CINZA = "595959"

FONTE_TITULO  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
FONTE_HEADER  = Font(name="Arial", bold=True, size=10, color="1F4E79")
FONTE_NORMAL  = Font(name="Arial", size=10)
FONTE_LABEL   = Font(name="Arial", bold=True, size=10)
FONTE_SUBTIT  = Font(name="Arial", bold=True, size=11, color="1F4E79")

BORDA_FINA = Border(
    left=Side(style="thin", color="B8B8B8"),
    right=Side(style="thin", color="B8B8B8"),
    top=Side(style="thin", color="B8B8B8"),
    bottom=Side(style="thin", color="B8B8B8"),
)
BORDA_MEDIA = Border(
    left=Side(style="medium", color="1F4E79"),
    right=Side(style="medium", color="1F4E79"),
    top=Side(style="medium", color="1F4E79"),
    bottom=Side(style="medium", color="1F4E79"),
)

def celula_fill(cor):
    return PatternFill("solid", start_color=cor, end_color=cor)


def aplicar_borda_range(ws, row_ini, col_ini, row_fim, col_fim, borda):
    for row in ws.iter_rows(min_row=row_ini, max_row=row_fim,
                             min_col=col_ini, max_col=col_fim):
        for cell in row:
            cell.border = borda
            
def criar_aba_info(workbook):
    worksheet = workbook.active
    worksheet.title = "Informações da Avaliação"
    worksheet.sheet_view.showGridLines = False
    
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 40
    worksheet.column_dimensions["C"].width = 18
    
    worksheet.row_dimensions[1].height = 32
    worksheet.merge_cells("A1:C1")
    cell = worksheet["A1"]
    cell.value = "Informações da Avaliação"
    cell.font = Font(name="Arial", bold=True, size=14, color=BRANCO)
    cell.fill = celula_fill(AZUL_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    worksheet.row_dimensions[2].height = 20
    worksheet.merge_cells("A2:C2")
    c = worksheet["A2"]
    c.value = "Preencha os campos abaixo antes de registrar as escalas nas abas correspondentes."
    c.font = Font(name="Arial", size=9, italic=True, color=CINZA)
    c.fill = celula_fill(AZUL_CLARO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    
    campos = [
        ("paciente_nome",          "Nome completo do paciente", False),
        ("data_nascimento",        "Data de nascimento",        False),
        ("fisioterapeuta_crefito", "CREFITO do fisioterapeuta", False),
        ("data_avaliacao",         "Data da avaliação",         False),
        ("observacoes",            "Observações (opcional)",    True),
    ]
    
    for i, (campo, label, opcional) in enumerate(campos, start=4):
        worksheet.row_dimensions[i].height = 22
        
        c_campo = worksheet.cell(row=i, column=1, value=campo)
        c_campo.font = FONTE_LABEL
        c_campo.fill = celula_fill(AZUL_CLARO)
        c_campo.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c_campo.border = BORDA_FINA
        
        c_label = worksheet.cell(row=i, column=2, value=label)
        c_label.font = FONTE_NORMAL
        c_label.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c_label.border = BORDA_FINA
        
        c_input = worksheet.cell(row=i, column=3, value="")
        c_input.font = FONTE_NORMAL
        c_input.fill = celula_fill(VERDE_INPUT if not opcional else AMARELO_OBS)
        c_input.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c_input.border = BORDA_MEDIA
        
        worksheet.row_dimensions[3].height = 14
        cell = worksheet.cell(row=3, column=1, value="Campo")
        cell.font = FONTE_HEADER
        cell.fill = celula_fill(AZUL_CLARO)
        cell.border = BORDA_FINA
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell = worksheet.cell(row=3, column=2, value="Descrição")
        cell.font = FONTE_HEADER
        cell.fill = celula_fill(AZUL_CLARO)
        cell.border = BORDA_FINA
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell = worksheet.cell(row=3, column=3, value="Valor")
        cell.font = FONTE_HEADER
        cell.fill = celula_fill(AZUL_CLARO)
        cell.border = BORDA_FINA
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
def criar_aba_escala(workbook, escala):
    worksheet = workbook.create_sheet(title=escala["nome_aba"])
    worksheet.sheet_view.showGridLines = False

    worksheet.column_dimensions["A"].width = 10
    worksheet.column_dimensions["B"].width = 58
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 14

    worksheet.row_dimensions[1].height = 30
    worksheet.merge_cells("A1:D1")
    cell = worksheet["A1"]
    cell.value = escala["nome"]
    cell.font = Font(name="Arial", bold=True, size=13, color=BRANCO)
    cell.fill = celula_fill(AZUL_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.row_dimensions[2].height = 18
    worksheet.merge_cells("A2:D2")
    cell = worksheet["A2"]
    cell.value = escala["descricao"]
    cell.font = Font(name="Arial", size=9, italic=True, color=CINZA)
    cell.fill = celula_fill(AZUL_CLARO)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.row_dimensions[3].height = 16
    for col, titulo in enumerate(["Nº", "Descrição do item", "Pontuação", "Máximo"], start=1):
        cell = worksheet.cell(row=3, column=col, value=titulo)
        cell.font = FONTE_HEADER
        cell.fill = celula_fill(AZUL_CLARO)
        cell.border = BORDA_FINA
        cell.alignment = Alignment(horizontal="center", vertical="center")

    primeira_linha_item = 4
    for i, (num, desc, maximo) in enumerate(escala["itens"]):
        row = primeira_linha_item + i
        worksheet.row_dimensions[row].height = 18
        cor_linha = CINZA_LINHA if i % 2 == 0 else BRANCO

        c_num = worksheet.cell(row=row, column=1, value=num)
        c_num.font = FONTE_NORMAL
        c_num.fill = celula_fill(cor_linha)
        c_num.alignment = Alignment(horizontal="center", vertical="center")
        c_num.border = BORDA_FINA

        c_desc = worksheet.cell(row=row, column=2, value=desc)
        c_desc.font = FONTE_NORMAL
        c_desc.fill = celula_fill(cor_linha)
        c_desc.alignment = Alignment(vertical="center", indent=1)
        c_desc.border = BORDA_FINA

        c_pont = worksheet.cell(row=row, column=3)
        c_pont.fill = celula_fill(VERDE_INPUT)
        c_pont.alignment = Alignment(horizontal="center", vertical="center")
        c_pont.border = BORDA_MEDIA
        c_pont.font = Font(name="Arial", size=10, bold=True)

        c_max = worksheet.cell(row=row, column=4,
                        value=maximo if maximo is not None else "—")
        c_max.font = Font(name="Arial", size=10, color=CINZA)
        c_max.fill = celula_fill(cor_linha)
        c_max.alignment = Alignment(horizontal="center", vertical="center")
        c_max.border = BORDA_FINA

    ultima_linha_item = primeira_linha_item + len(escala["itens"]) - 1
    linha_total = ultima_linha_item + 2

    worksheet.row_dimensions[linha_total].height = 22

    c_label = worksheet.cell(row=linha_total, column=1, value="TOTAL")
    c_label.font = Font(name="Arial", bold=True, size=11, color=BRANCO)
    c_label.fill = celula_fill(AZUL_HEADER)
    c_label.alignment = Alignment(horizontal="center", vertical="center")
    c_label.border = BORDA_FINA
    worksheet.merge_cells(f"A{linha_total}:B{linha_total}")

    c_total = worksheet.cell(row=linha_total, column=3)
    c_total.fill = celula_fill(AZUL_CLARO)
    c_total.alignment = Alignment(horizontal="center", vertical="center")
    c_total.border = BORDA_MEDIA
    c_total.font = Font(name="Arial", bold=True, size=11, color=AZUL_HEADER)

    if escala.get("formula_total"):
        inicio = primeira_linha_item
        fim = ultima_linha_item
        c_total.value = f"=SUM(C{inicio}:C{fim})"
    else:
        c_total.value = f"=C{primeira_linha_item}"

    worksheet.freeze_panes = "A4"
        

def gerar_template_avaliacao(caminho_saida: str):
    """Gera o template Excel de avaliação e salva no caminho especificado."""
    workbook = Workbook()
    criar_aba_info(workbook)
    
    for escala in get_escalas():
        criar_aba_escala(workbook, escala)
    workbook.save(caminho_saida)
    print(f"Template de avaliação gerado com sucesso em: {caminho_saida}")
    
if __name__ == "__main__":
    import os
    os.makedirs("data", exist_ok=True)
    gerar_template_avaliacao("data/template_avaliacao.xlsx")
    